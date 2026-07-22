from __future__ import annotations

import json
import re
from dataclasses import asdict, is_dataclass
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook

from src.models import SourceBlogger


INSTAGRAM_URL_RE = re.compile(r"https?://(?:www\.)?instagram\.com/([^/?#]+)/?", re.IGNORECASE)


def load_source_bloggers(xlsx_path: Path) -> list[SourceBlogger]:
    """Читает таблицу и нормализует строки блогеров."""
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    worksheet = workbook.worksheets[0]

    bloggers: list[SourceBlogger] = []

    for excel_row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        source_label = row[0] if len(row) > 0 else None
        raw_value = row[1] if len(row) > 1 else None

        if not source_label and not raw_value:
            continue

        normalized_value = str(raw_value).strip() if raw_value else ""
        url = extract_instagram_url(normalized_value)
        username = extract_username(normalized_value)
        status = "ok" if url and username else "needs_review"

        bloggers.append(
            SourceBlogger(
                row_number=excel_row_index,
                source_label=str(source_label) if source_label is not None else None,
                raw_value=normalized_value,
                url=url,
                username=username,
                status=status,
            )
        )

    return bloggers


def extract_instagram_url(raw_value: str) -> str | None:
    """Извлекает ссылку на Instagram-профиль."""
    match = INSTAGRAM_URL_RE.search(raw_value)
    if match:
        parsed = urlparse(match.group(0))
        return f"{parsed.scheme}://{parsed.netloc}/{match.group(1)}"

    username = extract_username(raw_value)
    if username:
        return f"https://www.instagram.com/{username}"

    return None


def extract_username(raw_value: str) -> str | None:
    """Извлекает username из ссылки или текста."""
    url_match = INSTAGRAM_URL_RE.search(raw_value)
    if url_match:
        return url_match.group(1)

    handle_match = re.search(r"@([A-Za-z0-9._]+)", raw_value)
    if handle_match:
        return handle_match.group(1)

    return None


def summarize_source_bloggers(bloggers: list[SourceBlogger]) -> dict[str, int]:
    """Считает краткую сводку по загруженным блогерам."""
    return {
        "total": len(bloggers),
        "ok": sum(1 for item in bloggers if item.status == "ok"),
        "needs_review": sum(1 for item in bloggers if item.status == "needs_review"),
    }


def get_problem_bloggers(bloggers: list[SourceBlogger]) -> list[SourceBlogger]:
    """Возвращает строки, требующие ручной проверки."""
    return [item for item in bloggers if item.status == "needs_review"]


def save_bloggers_to_json(bloggers: list[object], output_path: Path) -> None:
    """Сохраняет список объектов в JSON-файл."""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    serializable_items = []
    for item in bloggers:
        if is_dataclass(item):
            serializable_items.append(asdict(item))
        else:
            serializable_items.append(item)

    with output_path.open("w", encoding="utf-8") as file:
        json.dump(serializable_items, file, ensure_ascii=False, indent=2)
